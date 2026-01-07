import os
from openpyxl import load_workbook

from app import create_app
from app.extensions import db
from app.rent.models import Tool


def to_int_mnt(value) -> int:
    """
    Excel values can be floats (sometimes shown like 1.117692e+05).
    We round and convert to int MNT.
    """
    if value is None:
        return 0
    try:
        return int(round(float(value)))
    except Exception:
        s = str(value).replace(",", "").strip()
        try:
            return int(round(float(s)))
        except Exception:
            return 0


def seed_tools_from_excel(xlsx_path: str):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Read header row
    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip() if cell.value is not None else "")

    def col_index(name: str) -> int:
        # Return 0-based index
        for i, h in enumerate(header):
            if h.strip().lower() == name.strip().lower():
                return i
        return -1

    idx_part = col_index("Part number")
    idx_desc = col_index("Description")
    idx_qty = col_index("Available QTY")
    idx_price_1_7 = col_index("Daily price (1-7 days)")
    idx_price_8_30 = col_index("Daily price (8-30days)")

    if idx_desc == -1 or idx_qty == -1 or idx_price_1_7 == -1:
        raise ValueError("Missing required columns. Check Excel headers (Description, Available QTY, Daily price (1-7 days)).")

    created = 0
    updated = 0

    # Start reading from row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        part_number = (str(row[idx_part]).strip() if idx_part != -1 and row[idx_part] is not None else None)
        description = (str(row[idx_desc]).strip() if row[idx_desc] is not None else "")
        available_qty = int(row[idx_qty]) if row[idx_qty] is not None else 0

        price_1_7 = to_int_mnt(row[idx_price_1_7])
        price_8_30 = to_int_mnt(row[idx_price_8_30]) if idx_price_8_30 != -1 else 0

        if not description:
            continue
        if available_qty < 0:
            available_qty = 0
        if price_1_7 <= 0:
            continue

        # Name = first part of description (keep it simple)
        name = description.split(" (")[0].strip()
        if len(name) > 160:
            name = name[:160]

        # Upsert by part_number when available, else by name+price
        tool = None
        if part_number:
            tool = Tool.query.filter_by(part_number=part_number).first()

        if tool is None:
            tool = Tool.query.filter_by(name=name, daily_price=price_1_7).first()

        if tool:
            tool.part_number = part_number
            tool.name = name
            tool.description = description
            tool.available_qty = available_qty
            tool.daily_price = price_1_7
            tool.daily_price_8_30 = (price_8_30 if price_8_30 > 0 else None)
            updated += 1
        else:
            db.session.add(
                Tool(
                    part_number=part_number,
                    name=name,
                    description=description,
                    available_qty=available_qty,
                    daily_price=price_1_7,
                    daily_price_8_30=(price_8_30 if price_8_30 > 0 else None),
                )
            )
            created += 1

    db.session.commit()
    return created, updated


if __name__ == "__main__":
    app = create_app()
    with app.app_context():
        xlsx_path = os.path.join("data", "tool_rent_catalog.xlsx")
        created, updated = seed_tools_from_excel(xlsx_path)
        print(f"✅ Tools seeded. Created: {created}, Updated: {updated}")
